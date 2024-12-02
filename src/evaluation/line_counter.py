# Counts line from the file and add count to the excel new column

import os
import glob
import re
import pandas as pd
from openpyxl import load_workbook, Workbook


from openpyxl import load_workbook
import pandas as pd


def update_existing_column_excel(file_path, column_name, column_data):
    """
    Updates an existing empty column in an Excel file with new data.

    Args:
        file_path (str): Path to the Excel file.
        column_name (str): Name of the column to be updated.
        column_data (list): Data for the column, must match the number of rows in the sheet.
    """
    # Load the workbook and identify the active sheet
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Read the existing data into a DataFrame
    existing_data = pd.DataFrame(sheet.values)
    # Set first row as column headers
    existing_data.columns = existing_data.iloc[0]
    existing_data = existing_data[1:]  # Remove the header row from the data

    # Ensure the column exists in the file
    if column_name not in existing_data.columns:
        raise ValueError(
            f"The column '{column_name}' does not exist in the Excel file.")

    # Update the empty column with new data
    if len(column_data) != len(existing_data):
        raise ValueError(
            "Length of column_data must match the number of rows in the existing data.")
    existing_data[column_name] = column_data

    # Save the updated DataFrame back to the Excel file
    with pd.ExcelWriter(file_path, engine="openpyxl", mode="w") as writer:
        existing_data.to_excel(writer, index=False, startrow=1)

    print(f"Column '{column_name}' updated successfully.")


# Path to the folder containing the Python files
python_files = glob.glob(os.path.join(
    r"LLM-Improved-Code-Quality\test_data\temp\code\python", '*.py'))

# Path to the folder containing the Java files
java_files = glob.glob(os.path.join(
    r"LLM-Improved-Code-Quality\test_data\temp\code\java", '*.java'))

code_lines = []

for file_path in python_files:
    with open(file_path, 'r') as file:
        code = file.read()

        if code.startswith("Python"):
            # Remove the first line of the code
            code = code.split("\n", 1)[1]

        code_lines.append(code.count("\n") + 1)

for file_path in java_files:
    with open(file_path, 'r') as file:
        code = file.read()

        if code.startswith("Python"):
            # Remove the first line of the code
            code = code.split("\n", 1)[1]

        code_lines.append(code.count("\n") + 1)


print(len(code_lines))

# Example to update the 'lines_of_code' column
file_path = "LLM-Improved-Code-Quality/test_data/temp/submissions/submission_results.xlsx"
column_name = "lines_of_code"
column_data = code_lines  # New data to be added to the existing column

update_existing_column_excel(file_path, column_name, column_data)
