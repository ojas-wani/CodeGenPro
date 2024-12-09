import glob
from io import BytesIO
import os
import re
import string
import time
import logging
import undetected_chromedriver as uc
import pyperclip
import platform
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.core.os_manager import ChromeType
from webdriver_manager.firefox import GeckoDriverManager
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, NoSuchElementException

from openpyxl import load_workbook, Workbook
import pandas as pd
from PIL import Image


class AutoLeetCode:
    def __init__(self,
                 driver_executable_path: str = None,
                 driver_arguments: str = None,
                 headless: bool = False,
                 username: str = None,
                 password: str = None,
                 url: str = None,
                 verbose: bool = False,
                 incognito: bool = False,
                 skip_login: bool = False,
                 user_data_dir: str = None,
                 login_type: str = 'normal'):

        self.headless = headless

        if not skip_login:
            if not username or not password:
                logging.warning(
                    "The username and password parameters are deprecated and will be removed soon."
                    " Please adjust your environment variables to pass username and password."
                )
                raise NameError(
                    f'Either provide username or password.')
            username = username
            password = password

        self.url = url or "https://leetcode.com/accounts/login"
        self.task_url = "https://leetcode.com/problems/"
        if verbose:
            logging.getLogger().setLevel(logging.INFO)
            logging.info('Verbose mode active')

        options = uc.ChromeOptions()

        profile_path = r'C:\Users\wanio\AppData\Local\Google\Chrome\User Data'
        options.add_argument(
            f"--user-data-dir={profile_path}")
        options.add_argument("--profile-directory=Test")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")

        options.headless = self.headless

        if incognito:
            options.add_argument('--incognito')
        if driver_arguments:
            for _arg in driver_arguments:
                options.add_argument(_arg)

        logging.info('Loading undetected Chrome')
        self.browser = uc.Chrome(
            user_data_dir=user_data_dir,
            driver_executable_path=driver_executable_path,
            options=options,
            headless=headless,
            log_level=10,
            advanced_elements=True
        )

        agent = self.browser.execute_script("return navigator.userAgent")
        self.browser.execute_cdp_cmd(
            'Network.setUserAgentOverride',
            {"userAgent": agent.replace('Headless', '')}
        )
        self.browser.set_page_load_timeout(15)

        logging.info('Loaded undetected Chrome')
        logging.info('Opening LeetCode')
        self.browser.get(self.url)

        if not skip_login and login_type == 'normal':
            self.login(username, password)
        elif not skip_login and login_type == 'manully':
            self.login_manully(username, password)

    def login(self, username, password):
        """
        Performs the login process with the provided username and password.

        This function operates on the login page.
        It fills in the username and password fields. 
        It finds and clicks the login button 
        """
        time.sleep(2)
        # Find the email textbox and fill it with the username
        email_box = self.sleepy_find_element(By.ID, 'id_login')
        email_box.send_keys(username)
        logging.info('Email box filled')

        # Find the password textbox and fill it with the password
        password_box = self.sleepy_find_element(By.ID, 'id_password')
        password_box.send_keys(password)
        logging.info('Password box filled')

        # Find the login button and click it
        login_button = self.sleepy_find_element(By.ID, 'signin_btn')
        login_button.click()
        logging.info('Login button clicked')
        time.sleep(1)

    def login_manully(self, username, password, timeout: int = 15):
        if "login" in self.browser.current_url:
            time.sleep(2)
            # Find the email textbox and fill it with the username
            email_box = self.sleepy_find_element(By.ID, 'id_login')
            email_box.send_keys(username)
            logging.info('Email box filled')

            # Find the password textbox and fill it with the password
            password_box = self.sleepy_find_element(By.ID, 'id_password')
            password_box.send_keys(password)
            logging.info('Password box filled')

        try:
            WebDriverWait(self.browser, timeout).until(
                lambda driver: self.check_login_success_condition()
            )
            logging.info('Login process is completed manually')
        except TimeoutError:
            logging.error('Login process is not completed manually')

    def check_login_success_condition(self):
        return "login" not in self.browser.current_url and "error" not in self.browser.current_url and "leetcode.com" in self.browser.current_url

    def run_and_collect(self, task_info: dict, language: str, code: str):
        """
        Run the code to the problem.

        This function operates on the problem page.
        It fills in the code editor with the provided code.
        It finds and clicks the submit button.
        It waits for the submission to complete.
        """
        logging.info('Running the code')
        task_url = self.task_url + task_info["task_name"]
        self.browser.get(task_url)
        logging.info('Opening %s', task_url)
        time.sleep(3)

        return_result_status, return_result_details = "", ""

        # Check if need to change language
        try:
            code_editor = WebDriverWait(self.browser, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "div[data-keybinding-context='1']"))
            )
        except TimeoutException:
            logging.error('Timed out waiting for code editor to load')
            # Capture a screenshot for debugging
            self.browser.save_screenshot('timeout_code_editor.png')
            return return_result_status, return_result_details

        current_language = code_editor.get_attribute("data-mode-id")

        if current_language != language.lower():
            # Find the language dropdown and click it
            language_dropdown_button_selector = "button.rounded.items-center.whitespace-nowrap.focus\:outline-none.inline-flex.bg-transparent"
            try:
                language_dropdown_button = WebDriverWait(self.browser, 10).until(
                    EC.element_to_be_clickable(
                        (By.CSS_SELECTOR, language_dropdown_button_selector))
                )
                language_dropdown_button.click()
            except TimeoutException:
                logging.error('Timed out waiting for language dropdown button')
                self.browser.save_screenshot('timeout_language_dropdown.png')
                return return_result_status, return_result_details

            # Find the language option and click it
            csharp_option_xpath = "//div[contains(text(), '" + language + "')]"
            try:
                csharp_option = WebDriverWait(self.browser, 3).until(
                    EC.element_to_be_clickable((By.XPATH, csharp_option_xpath))
                )
                csharp_option.click()
            except TimeoutException:
                logging.error('Timed out waiting for language option to load')
                self.browser.save_screenshot('timeout_language_option.png')
                return return_result_status, return_result_details

        # Find the code editor and fill it with the code
        code_textarea_selector = "textarea.inputarea"
        try:
            code_textarea = WebDriverWait(self.browser, 10).until(
                EC.visibility_of_element_located(
                    (By.CSS_SELECTOR, code_textarea_selector))
            )
            is_focused = False

            while not is_focused:
                code_textarea.click()

                # Check if code_textarea is focused
                is_focused = code_textarea.is_enabled()

            print("@#%$#^$&@$^$%&$^%$ Is focused: ", is_focused)

        except TimeoutException:
            logging.error('Timed out waiting for code textarea')
            self.browser.save_screenshot('timeout_code_textarea.png')
            return return_result_status, return_result_details

        # Use ActionChains to select all text
        actions = ActionChains(self.browser)
        # For Windows/Linux, use CTRL. For Mac, use COMMAND.
        modifier_key = Keys.COMMAND if platform.system() == 'Darwin' else Keys.CONTROL
        actions.key_down(modifier_key).send_keys(
            'a').key_up(modifier_key).perform()
        # Wait briefly to ensure the action is completed
        time.sleep(0.5)

        # Delete the selected text
        actions = ActionChains(self.browser)
        actions.send_keys(Keys.DELETE).perform()

        # Copy your new code to clipboard
        pyperclip.copy(code)
        # Paste the new code into the editor
        actions = ActionChains(self.browser)
        actions.key_down(modifier_key).send_keys(
            'v').key_up(modifier_key).perform()
        time.sleep(1)

        # Find the submit button and click it
        submit_button_selector = "button[data-e2e-locator='console-run-button']"
        try:
            submit_button = WebDriverWait(self.browser, 10).until(
                EC.element_to_be_clickable(
                    (By.CSS_SELECTOR, submit_button_selector))
            )
            submit_button.click()
        except TimeoutException:
            logging.error('Timed out waiting for submit button')
            self.browser.save_screenshot('timeout_submit_button.png')
            return return_result_status, return_result_details

        # Wait for the submission to complete
        logging.info('Waiting for submission to complete')
        try:
            result_element = WebDriverWait(self.browser, 30).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "div[data-e2e-locator='console-result'], span[data-e2e-locator='console-result']"))
            )
            result_text = result_element.text
            logging.info('Submission result: %s', result_text)
            return_result_status = result_text

            if "Error" in result_text:
                try:
                    # Using a more general CSS selector for error details
                    runtime_error_elements = self.browser.find_elements(
                        By.CSS_SELECTOR, ".font-menlo.text-red-3.dark\\:text-dark-red-3.whitespace-pre-wrap.break-all.text-xs")
                    if runtime_error_elements:
                        return_result_details = runtime_error_elements[0].text
                    else:
                        logging.error('No runtime error details found')
                except NoSuchElementException:
                    logging.error('Error details not found')

        except TimeoutException:
            logging.error('Timed out waiting for submission result')

            # Capture a screenshot for debugging
            self.browser.save_screenshot(
                f"LLM-Improved-Code-Quality/test_data/temp2/updated_code/submissions/submission-errors/{task_info['id']}_timeout_submission_result.png")

        return return_result_status, return_result_details

    def submit_and_collect(self, task_info: dict, language: str, code: str):
        """
        Submits the code to the problem.

        This function operates on the problem page.
        It fills in the code editor with the provided code.
        It finds and clicks the submit button.
        It waits for the submission to complete.
        """
        logging.info('Submitting code')
        # task_url = self.task_url + problem_id
        task_url = self.task_url + task_info["task_name"]
        self.browser.get(task_url)
        logging.info('Opening %s', task_url)
        time.sleep(1)

        # check if need change language
        code_editor = WebDriverWait(self.browser, 10).until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, "div[data-keybinding-context='1']"))
        )
        current_language = code_editor.get_attribute("data-mode-id")

        difficulty_element = self.browser.find_elements(
            By.XPATH, "//div[contains(@class, 'text-difficulty-easy') or contains(@class, 'text-difficulty-medium') or contains(@class, 'text-difficulty-hard')]")
        # Extract the difficulty text
        difficulty = "Not Found"
        if difficulty_element:
            difficulty = difficulty_element[0].text
        logging.info('Difficulty level: %s', difficulty)

        if current_language != language.lower():

            # Find the language dropdown and click it
            language_dropdown_button_selector = "button.rounded.items-center.whitespace-nowrap.focus\:outline-none.inline-flex.bg-transparent"
            language_dropdown_button = WebDriverWait(self.browser, 30).until(
                EC.element_to_be_clickable(
                    (By.CSS_SELECTOR, language_dropdown_button_selector))
            )
            language_dropdown_button.click()

            # Find the language dropdown and click it
            csharp_option_xpath = "//div[contains(text(), '"+language+"')]"
            csharp_option = WebDriverWait(self.browser, 3).until(
                EC.element_to_be_clickable((By.XPATH, csharp_option_xpath))
            )
            csharp_option.click()

        # Find the code editor and fill it with the code
        code_textarea_selector = "textarea.inputarea"
        try:
            code_textarea = WebDriverWait(self.browser, 10).until(
                EC.visibility_of_element_located(
                    (By.CSS_SELECTOR, code_textarea_selector))
            )

            is_focused = False

            while not is_focused:
                code_textarea.click()

                # Check if code_textarea is focused
                is_focused = code_textarea.is_enabled()

            print("@#%$#^$&@$^$%&$^%$ Is focused: ", is_focused)

        except TimeoutException:
            logging.error('Timed out waiting for code textarea')
            self.browser.save_screenshot('timeout_code_textarea.png')

        # Use ActionChains to select all text
        actions = ActionChains(self.browser)
        # For Windows/Linux, use CTRL. For Mac, use COMMAND.
        modifier_key = Keys.COMMAND if platform.system() == 'Darwin' else Keys.CONTROL
        actions.key_down(modifier_key).send_keys(
            'a').key_up(modifier_key).perform()
        # Wait briefly to ensure the action is completed
        time.sleep(0.5)

        # Delete the selected text
        actions = ActionChains(self.browser)
        actions.send_keys(Keys.DELETE).perform()

        # Copy your new code to clipboard
        pyperclip.copy(code)
        # Paste the new code into the editor
        actions = ActionChains(self.browser)
        actions.key_down(modifier_key).send_keys(
            'v').key_up(modifier_key).perform()
        time.sleep(1)

        # Find the submit button and click it
        submit_button_selector = "button[data-e2e-locator='console-submit-button']"
        submit_button = self.sleepy_find_element(
            By.CSS_SELECTOR, submit_button_selector)

        # submit_button = WebDriverWait(self.browser, 10).until(
        #     EC.element_to_be_clickable(
        #         (By.CSS_SELECTOR, submit_button_selector))
        # )
        while EC.url_changes(self.browser.current_url):
            try:
                submit_button.click()
                WebDriverWait(self.browser, 10, 1, ignored_exceptions=TimeoutException).until(
                    EC.url_changes(self.browser.current_url))
                break
            except:
                print("--------------------Error in clicking submit button")
        # submit_button.click()

        # Wait for the submission to complete
        logging.info('Waiting for submission to complete')
        try:
            result_element = WebDriverWait(self.browser, 20).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "span[data-e2e-locator='submission-result']"))
            )

            result_text = result_element.text

            logging.info('Submission result: %s', result_text)

        except TimeoutException:  # if successful submission not found

            logging.error('Timed out waiting for submission result')

            try:
                # Check for Wrong Answer and how many test cases failed
                wrong_answer_element = self.browser.find_elements(
                    By.CSS_SELECTOR, "h3.flex.items-center.text-xl.font-semibold.text-red-60.dark\\:text-red-60")

                if wrong_answer_element:
                    result_text = wrong_answer_element[0].text.splitlines(
                    )[0] + " " + wrong_answer_element[0].text.splitlines()[1]
                    logging.error('Wrong Answer: %s',
                                  wrong_answer_element[0].text)

                    wrong_answer_div = self.browser.find_element(
                        By.CSS_SELECTOR, 'div.mx-auto.flex.w-full.max-w-\[700px\].flex-col.gap-4.px-4.py-3')

                    # wrong_answer_div = self.browser.find_element(
                    #     By.CSS_SELECTOR, 'div.space-y-4')

                    # x = int(wrong_answer_div.get_attribute("offsetLeft"))
                    # y = int(wrong_answer_div.get_attribute("offsetTop"))
                    # w = int(wrong_answer_div.get_attribute("offsetWidth"))
                    # h = int(wrong_answer_div.get_attribute("offsetHeight"))

                    # screenshot_png = self.browser.get_screenshot_as_png()
                    # screenshot = Image.open(BytesIO(screenshot_png))
                    # cropped_screenshot = screenshot.crop((x, y, x+w, y+h))

                    # cropped_screenshot.save(
                    #     f"LLM-Improved-Code-Quality/test_data/temp2/updated_code/submissions/code1_submission-errors/python/wrong_answer/{task_info['id']}-{task_info['task_name']}.png")

                    # self.browser.get_screenshot_as_png(
                    #     f"LLM-Improved-Code-Quality/test_data/temp2/updated_code/submissions/code1_submission-errors/python/wrong_answer/{task_info['id']}-{task_info['task_name']}.png")

                    wrong_answer_div.screenshot(
                        f"LLM-Improved-Code-Quality/test_data/temp2/updated_code/submissions/code1_submission-errors/python/wrong_answer/{task_info['id']}-{task_info['task_name']}.png")

                else:
                    # Check for Runtime Error
                    runtime_error_elements = WebDriverWait(self.browser, 10).until(
                        EC.presence_of_all_elements_located(
                            (By.CSS_SELECTOR, "div.flex.w-full.flex-wrap.items-center"))
                    )

                    if runtime_error_elements:
                        result_text = "Runtime Error"
                        logging.error('Runtime Error: %s',
                                      runtime_error_elements[0].text)

                        error_div = self.browser.find_element(
                            By.CSS_SELECTOR, 'div.mx-auto.flex.w-full.max-w-\[700px\].flex-col.gap-4.px-4.py-3')

                        error_div.screenshot(
                            f"LLM-Improved-Code-Quality/test_data/temp2/updated_code/submissions/code1_submission-errors/python/runtime_error/{task_info['id']}-{task_info['task_name']}.png")

                        # self.browser.save_screenshot(
                        #     f"LLM-Improved-Code-Quality/test_data/temp2/updated_code/submissions/code1_submission-errors/python/runtime_error/{task_info['id']}-{task_info['task_name']}.png")
                    else:
                        result_text = "Timeout"

            except NoSuchElementException or TimeoutException:
                logging.error('Error details not found')

        # Append the result status to the task info and submission result to the excel file
        task_info["difficulty"] = difficulty
        task_info["submission_result"] = result_text

        # if language == "Python3":
        self.append_row_excel(
            r"LLM-Improved-Code-Quality\test_data\temp2\updated_code\submissions\code1_submission_results.xlsx", task_info)

    def sleepy_find_element(self, by: By, query: str, attempt_count: int = 20, sleep_duration: int = 1):
        '''
        Finds the web element using the locator and query.

        This function attempts to find the element multiple times with a specified
        sleep duration between attempts. If the element is found, the function returns the element.

        Args:
            by (selenium.webdriver.common.by.By): The method used to locate the element.
            query (str): The query string to locate the element.
            attempt_count (int, optional): The number of attempts to find the element. Default: 20.
            sleep_duration (int, optional): The duration to sleep between attempts. Default: 1.

        Returns:
            selenium.webdriver.remote.webelement.WebElement: Web element or None if not found.
        '''
        for _count in range(attempt_count):
            item = self.browser.find_elements(by, query)
            if len(item) > 0:
                item = item[0]
                logging.info('Element %s has found', query)
                break
            logging.info('Element %s is not present, attempt: %d',
                         query, _count+1)
            time.sleep(sleep_duration)
        return item

    def append_row_excel(self, file_path, row_data):

        # Convert the row data to a DataFrame with a single row
        new_row_df = pd.DataFrame([row_data])

        # Load the workbook and identify the sheet
        workbook = load_workbook(file_path)
        sheet = workbook.active
        startrow = sheet.max_row  # Get the starting row to append data

        # Append the new row
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists="overlay") as writer:
            new_row_df.to_excel(writer, index=False,
                                header=False, startrow=startrow)

        print("Row appended successfully.")


if __name__ == '__main__':
    autoleet = AutoLeetCode(
        headless=False,
        username="mrflametest",
        password="Mrflame@203",
        verbose=True,
        incognito=False,
        skip_login=True,
        # user_data_dir="User data/Test/",
        login_type='normal'
    )
    time.sleep(2)

    # Path to the folder containing the Python files
    python_files = glob.glob(os.path.join(
        r"LLM-Improved-Code-Quality\test_data\temp2\updated_code\code\python", '*.py'))
    # java_files = glob.glob(os.path.join(
    #     r"LLM-Improved-Code-Quality\test_data\temp\code\java", '*.java'))
    java_files = glob.glob(os.path.join(
        r"LLM-Improved-Code-Quality\test_data\temp2\updated_code\code1\java", '*.java'))

    for file_path in python_files:
        with open(file_path, 'r') as file:
            code = file.read()

            if code.startswith("Python"):
                # Remove the first line of the code
                code = code.split("\n", 1)[1]

            file_name = file.name.split("\\")[-1]
            problem_id = file_name.split("-")[0]
            task_name = re.search(r'\d+-(.*)\.py', file_name)

            llm_generated_code = code
            language = "Python3"
            # language = "Java"

            task_info = {
                "id": problem_id,
                "task_name": task_name.group(1),
                "difficulty": "",
                "language": language,
                "submission_result": ""
            }

            # sleep for 3 seconds before submitting the code to avoid too-soon submission leetcode error
            time.sleep(3)

            # autoleet.run_and_collect(task_info, language, llm_generated_code)

            # if submit codd and collect
            autoleet.submit_and_collect(
                task_info, language, llm_generated_code)

            break

    # Add name, id, and language here into the excel file here and add the result status above.
    # Put all into the pandas dataframe and append to the excel file.

    # if only run code and collect
    # result_status, result_details = autoleet.run_and_collect(
    #     task_name, language, chatgpt_generated_code)
